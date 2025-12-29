from __future__ import annotations

from dataclasses import dataclass
from os import PathLike
from pathlib import Path
from typing import Literal, TypedDict

import pandas as pd

from app.core.common.columns import canonicalize_headers
from app.infra.io_utils import ALT_CODE_COLUMN

PoolType = Literal["inspactor", "matrix"]


class SheetEvidence(TypedDict):
    sheet: str
    missing_columns: list[str]
    missing_count: int
    row_count: int | None
    has_mentor_id: bool
    excluded: bool
    exclusion_reason: str | None


@dataclass(frozen=True)
class PoolDetectionResult:
    pool_type: PoolType
    selected_sheet: str
    detection_method: str
    confidence: float
    evidence: dict[str, object]


_EXPECTED_COLUMNS: dict[PoolType, set[str]] = {
    "inspactor": {
        "نام معلم",
        "نام مدیر",
        "کد معلم",
        "کدپستی",
        "تعداد مدارس تحت پوشش",
        "کد ملی معلم جایگزین",
        "ظرفیت ویژه",
    },
    "matrix": set(),
}


def detect_pool_sheet(
    path: Path | str | PathLike[str],
    pool_type: PoolType,
    explicit_sheet: str | None = None,
) -> PoolDetectionResult:
    path = Path(path)
    excel = pd.ExcelFile(path)
    workbook = None
    workbook_opened = False
    if path.exists():
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=True)
            workbook_opened = True
        except ImportError:
            workbook = getattr(excel, "book", None)
    else:
        workbook = getattr(excel, "book", None)
    sheet_names = list(excel.sheet_names)

    if not sheet_names:
        raise ValueError(f"Workbook {path} has no sheets")

    if explicit_sheet and explicit_sheet not in sheet_names:
        raise ValueError(
            f"Requested sheet '{explicit_sheet}' not found in workbook; available: {sheet_names}"
        )

    evidence: list[SheetEvidence] = []
    expected = _EXPECTED_COLUMNS[pool_type]

    for sheet in sheet_names:
        header_frame = excel.parse(sheet, nrows=0)
        columns = {str(column).strip() for column in header_frame.columns}
        missing_columns = sorted(col for col in expected if col not in columns)
        missing_columns_lite = missing_columns[:5]
        missing_count = len(missing_columns)
        has_mentor_id = "mentor_id" in columns or "کد معلم" in columns
        is_reserved_matrix = pool_type == "inspactor" and sheet == "matrix"
        if explicit_sheet is not None and sheet != explicit_sheet:
            exclusion_reason = "not_explicit_sheet"
        elif is_reserved_matrix and explicit_sheet != "matrix":
            exclusion_reason = "reserved_sheet_matrix"
        else:
            exclusion_reason = None
        row_count: int | None = None
        if workbook is not None and sheet in getattr(workbook, "sheetnames", []):
            max_row = getattr(workbook[sheet], "max_row", None)
            row_count = max(max_row - 1, 0) if max_row is not None else None

        evidence.append(
            {
                "sheet": sheet,
                "missing_columns": missing_columns_lite,
                "missing_count": missing_count,
                "row_count": row_count,
                "has_mentor_id": has_mentor_id,
                "excluded": exclusion_reason is not None,
                "exclusion_reason": exclusion_reason,
            }
        )

    selection_reason: dict[str, object] | None = None

    if explicit_sheet is not None:
        selected_sheet = explicit_sheet
        detection_method = "explicit_sheet"
        confidence = 1.0
    else:
        candidates = [info for info in evidence if not info["excluded"]]
        if not candidates:
            raise ValueError(
                "No usable sheets found after exclusions; 'matrix' is reserved for pool_type='matrix'. "
                "Pass explicit_sheet='matrix' or use pool_type='matrix'.",
            )

        best_missing_count = min(info["missing_count"] for info in candidates)
        best_candidates = [
            info for info in candidates if info["missing_count"] == best_missing_count
        ]
        best_candidates.sort(
            key=lambda info: (
                info["missing_count"],
                info["row_count"] is None,
                -info["row_count"] if info["row_count"] is not None else 0,
                info["sheet"],
            )
        )

        selected_sheet = best_candidates[0]["sheet"]
        detection_method = "best_header_match"
        confidence = 1.0 if best_missing_count == 0 else 0.8
        selection_reason = {
            "missing_required_count": best_missing_count,
            "sort_key": (
                best_missing_count,
                best_candidates[0]["row_count"] is None,
                -best_candidates[0]["row_count"]
                if best_candidates[0]["row_count"] is not None
                else 0,
                best_candidates[0]["sheet"],
            ),
        }

    if (
        pool_type == "inspactor"
        and explicit_sheet is None
        and selected_sheet != "matrix"
        and any(info["sheet"] == "matrix" for info in evidence)
    ):
        matrix_info = next(info for info in evidence if info["sheet"] == "matrix")
        selected_info = next(info for info in evidence if info["sheet"] == selected_sheet)
        matrix_rows = matrix_info.get("row_count")
        selected_rows = selected_info.get("row_count")
        if matrix_info["has_mentor_id"] and matrix_rows is not None and (
            selected_rows is None or matrix_rows > selected_rows
        ):
            selected_sheet = "matrix"
            detection_method = "fallback_matrix_preferred"
            confidence = 0.9
            selection_reason = {
                "fallback": "matrix_has_more_rows_and_mentor_id",
                "matrix_rows": matrix_rows,
                "selected_rows": selected_rows,
            }

    return PoolDetectionResult(
        pool_type=pool_type,
        selected_sheet=selected_sheet,
        detection_method=detection_method,
        confidence=confidence,
        evidence={
            "path": str(path),
            "sheets": evidence,
            **({"selection_reason": selection_reason} if selection_reason else {}),
        },
    )

    if workbook_opened and hasattr(workbook, "close"):
        workbook.close()


def load_pool(
    path: Path | str | PathLike[str],
    *,
    pool_type: PoolType = "inspactor",
    pool_sheet: str | None = None,
) -> pd.DataFrame:
    """Load mentor pool workbook using shared sheet selection logic."""

    detection = detect_pool_sheet(path, pool_type=pool_type, explicit_sheet=pool_sheet)
    with pd.ExcelFile(path) as workbook:
        df = workbook.parse(detection.selected_sheet)
    canonical = canonicalize_headers(df, header_mode="fa")
    if ALT_CODE_COLUMN in canonical.columns:
        canonical = canonical.copy()
        canonical[ALT_CODE_COLUMN] = canonical[ALT_CODE_COLUMN].astype(str)
    canonical.attrs["pool_detection"] = detection
    return canonical
