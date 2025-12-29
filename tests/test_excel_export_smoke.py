from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.core.allocate_students import allocate_batch
from app.core.policy_adapter import policy as policy_adapter
from app.infra.cli_legacy import _make_excel_safe
from app.infra.config_flags import UserSettings
from app.infra.excel import export_allocations
from app.infra.io_utils import write_xlsx_atomic


def _build_sample_frames() -> tuple[pd.DataFrame, pd.DataFrame, str]:
    capacity_column = policy_adapter.stage_column("capacity_gate")
    assert capacity_column is not None

    students = pd.DataFrame(
        [
            {
                "student_id": "STD-1",
                "کدرشته": 1,
                "گروه آزمایشی": "تجربی",
                "جنسیت": 1,
                "دانش آموز فارغ": 0,
                "مرکز گلستان صدرا": 0,
                "مالی حکمت بنیاد": 0,
                "کد مدرسه": 1010,
            },
            {
                "student_id": "STD-2",
                "کدرشته": 1,
                "گروه آزمایشی": "تجربی",
                "جنسیت": 1,
                "دانش آموز فارغ": 0,
                "مرکز گلستان صدرا": 0,
                "مالی حکمت بنیاد": 0,
                "کد مدرسه": 1010,
            },
        ]
    )

    candidate_pool = pd.DataFrame(
        [
            {
                "پشتیبان": "Mentor A",
                "کد کارمندی پشتیبان": "EMP-1",
                "کدرشته": 1,
                "گروه آزمایشی": "تجربی",
                "جنسیت": 1,
                "دانش آموز فارغ": 0,
                "مرکز گلستان صدرا": 0,
                "مالی حکمت بنیاد": 0,
                "کد مدرسه": 1010,
                capacity_column: 2,
                "allocations_new": 0,
                "mentor_sort_key": 1,
            }
        ]
    )

    return students, candidate_pool, capacity_column


def _build_workbook(settings: UserSettings, output: Path) -> Path:
    students, candidate_pool, capacity_column = _build_sample_frames()
    batch_result = allocate_batch(
        students,
        candidate_pool,
        policy=policy_adapter.config,
        capacity_column=capacity_column,
    )

    sheets: dict[str, pd.DataFrame] = {
        "allocations": _make_excel_safe(batch_result.allocations_df),
        "updated_pool": _make_excel_safe(batch_result.pool_output),
        "logs": _make_excel_safe(batch_result.logs_df),
    }

    if settings.enable_trace_export:
        sheets["trace"] = _make_excel_safe(batch_result.trace_df)

    if settings.enable_trace_debug_sheets:
        debug_sheets = export_allocations.collect_trace_debug_sheets(
            batch_result.trace_df,
            students_df=students,
            history_info_df=None,
            policy=policy_adapter.config,
            summary_df=batch_result.trace_extras.summary_df,
            unallocated_summary=batch_result.trace_extras.unallocated_summary,
            policy_violations=batch_result.trace_extras.policy_violations,
            final_status_counts=batch_result.trace_extras.final_status_counts,
            enable_history_metrics=settings.enable_history_metrics,
        )
        for name, df in debug_sheets.items():
            sheets[name] = _make_excel_safe(df)

    write_xlsx_atomic(sheets, output, header_mode=None)
    return output


def _assert_core_sheets_present(sheet_names: set[str]) -> None:
    expected = {"allocations", "updated_pool", "logs"}
    missing = expected - sheet_names
    assert not missing, f"Missing core sheets: {missing}"


def _assert_id_headers_present(header_values: set[str]) -> None:
    assert any("student_id" in value for value in header_values), "student_id header missing"
    assert any("mentor_id" in value for value in header_values), "mentor_id header missing"


def test_excel_export_core_and_diagnostic_sheets_off(tmp_path: Path) -> None:
    settings = UserSettings(
        enable_history_metrics=False,
        enable_trace_debug_sheets=False,
        enable_trace_export=False,
    )
    workbook_path = _build_workbook(settings, tmp_path / "export-off.xlsx")

    workbook = load_workbook(workbook_path)
    sheet_names = set(workbook.sheetnames)
    _assert_core_sheets_present(sheet_names)

    diagnostic_sheets = {
        "trace",
        "HistoryMetrics",
        "summary_df",
        "unallocated_summary",
        "policy_violations",
        "FinalStatus_counts",
        "JoinKeyProvenance_counts",
    }

    assert sheet_names.isdisjoint(diagnostic_sheets)

    allocations_headers = {str(cell.value) for cell in next(workbook["allocations"].iter_rows(max_row=1))}
    _assert_id_headers_present(allocations_headers)


def test_excel_export_core_and_diagnostic_sheets_on(tmp_path: Path) -> None:
    settings = UserSettings(
        enable_history_metrics=True,
        enable_trace_debug_sheets=True,
        enable_trace_export=True,
    )
    workbook_path = _build_workbook(settings, tmp_path / "export-on.xlsx")

    workbook = load_workbook(workbook_path)
    sheet_names = set(workbook.sheetnames)
    _assert_core_sheets_present(sheet_names)

    expected_diagnostics = {"trace", "HistoryMetrics"}
    assert expected_diagnostics.issubset(sheet_names)

    header_row = {str(cell.value) for cell in next(workbook["allocations"].iter_rows(max_row=1))}
    _assert_id_headers_present(header_row)
