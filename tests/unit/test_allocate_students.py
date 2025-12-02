from __future__ import annotations

import importlib
import json
from dataclasses import replace
from pathlib import Path

import pandas as pd
import pytest

from app.core.allocate_students import (
    _normalize_pool,
    _separate_school_students,
    allocate_batch,
    allocate_student,
    build_selection_reason_rows,
)
from app.core.canonical_frames import (
    POOL_DUPLICATE_SUMMARY_ATTR,
    POOL_JOIN_KEY_DUPLICATES_ATTR,
    _build_join_key_duplicate_report,
    canonicalize_allocation_frames,
    canonicalize_pool_frame,
    canonicalize_students_frame,
    sanitize_pool_for_allocation,
)
from app.core.common import columns
from app.core.common.reasons import ReasonCode
from app.core.common.types import CANONICAL_JOIN_KEYS, JoinKeyValues
from app.core.policy_loader import PolicyConfig, load_policy, parse_policy_dict
from app.infra.canonical_frames import build_student_group_crosswalk
from app.infra.excel_writer import write_selection_reasons_sheet


@pytest.fixture()
def _base_pool() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "پشتیبان": ["زهرا", "علی"],
            "کد کارمندی پشتیبان": ["EMP-001", "EMP-002"],
            "کدرشته": [27, 27],
            "کدرشته | group_code": [27, 27],
            "گروه آزمایشی": [27, 27],
            "جنسیت": [1, 1],
            "جنسیت | gender": [1, 1],
            "دانش آموز فارغ": [0, 0],
            "دانش آموز فارغ | graduation_status": [0, 0],
            "مرکز گلستان صدرا": [1, 1],
            "مرکز گلستان صدرا | center": [1, 1],
            "مالی حکمت بنیاد": [0, 0],
            "مالی حکمت بنیاد | finance": [0, 0],
            "کد مدرسه": [3581, 3581],
            "کد مدرسه | school_code": [3581, 3581],
            "remaining_capacity": [2, 2],
            "occupancy_ratio": [0.1, 0.2],
            "allocations_new": [0, 0],
        }
    )


def _single_student(**overrides: object) -> pd.DataFrame:
    base = {
        "student_id": "STD-001",
        "کدرشته": 27,
        "گروه_آزمایشی": 27,
        "جنسیت": 1,
        "دانش_آموز_فارغ": 0,
        "مرکز_گلستان_صدرا": 1,
        "مالی_حکمت_بنیاد": 0,
        "کد_مدرسه": 3581,
    }
    base.update(overrides)
    return pd.DataFrame([base])


def test_canonicalize_students_frame_infers_missing_exam_group() -> None:
    policy = load_policy()
    students = pd.DataFrame(
        {
            "student_id": ["STD-001"],
            "کدرشته": [27],
            "جنسیت": [1],
            "دانش آموز فارغ": [0],
            "مرکز گلستان صدرا": [1],
            "مالی حکمت بنیاد": [0],
            "کد مدرسه": [3581],
        }
    )

    normalized = canonicalize_students_frame(students, policy=policy)

    exam_group_col = columns.CANON_EN_TO_FA["exam_group"]
    assert exam_group_col in normalized.columns
    assert normalized[exam_group_col].isna().all()


def test_canonicalize_students_frame_handles_duplicate_school_columns() -> None:
    policy = load_policy()
    students = pd.DataFrame(
        {
            "student_id": ["STD-001"],
            "کدرشته": [27],
            "جنسیت": [1],
            "دانش آموز فارغ": [0],
            "مرکز گلستان صدرا": [1],
            "مالی حکمت بنیاد": [0],
            "مدرسه نهایی": ["1111"],
            "school final": ["9999"],
        }
    )

    normalized = canonicalize_students_frame(students, policy=policy)

    school_fa = columns.CANON_EN_TO_FA["school_code"]
    assert normalized.columns.tolist().count(school_fa) == 1
    assert normalized["school_code_raw"].iloc[0] == "1111"


def test_canonicalize_students_frame_promotes_final_exam_column_variants() -> None:
    policy = load_policy()
    students = pd.DataFrame(
        {
            "student_id": ["STD-001"],
            "گروه آزمایشی نهایی (کد رشته)": [27],
            "جنسیت": [1],
            "دانش آموز فارغ": [0],
            "مرکز گلستان صدرا": [1],
            "مالی حکمت بنیاد": [0],
            "کد مدرسه": [3581],
        }
    )

    normalized = canonicalize_students_frame(students, policy=policy)

    group_col = columns.CANON_EN_TO_FA["group_code"]
    assert group_col in normalized.columns
    assert normalized[group_col].iloc[0] == 27


def test_canonicalize_students_frame_flattens_multiindex_school_columns() -> None:
    policy = load_policy()
    multi_columns = pd.MultiIndex.from_tuples(
        [
            ("student_id", ""),
            ("کدرشته", ""),
            ("گروه آزمایشی", ""),
            ("جنسیت", ""),
            ("دانش آموز فارغ", ""),
            ("مرکز گلستان صدرا", ""),
            ("مالی حکمت بنیاد", ""),
            (columns.CANON_EN_TO_FA["school_code"], "اول"),
            (columns.CANON_EN_TO_FA["school_code"], "دوم"),
        ]
    )
    students = pd.DataFrame(
        [
            [
                "STD-001",
                27,
                27,
                1,
                0,
                1,
                0,
                "1357",
                "5799",
            ]
        ],
        columns=multi_columns,
    )

    normalized = canonicalize_students_frame(students, policy=policy)

    school_fa = columns.CANON_EN_TO_FA["school_code"]
    assert normalized["school_code_raw"].iloc[0] == "1357"
    assert normalized[school_fa].iloc[0] == 1357


def test_canonicalize_pool_frame_handles_duplicate_mentor_columns(
    _base_pool: pd.DataFrame,
) -> None:
    policy = load_policy()
    pool = _base_pool.copy()
    pool.insert(0, "mentor_dup", pool["کد کارمندی پشتیبان"])
    columns = pool.columns.tolist()
    columns[0] = "کد کارمندی پشتیبان"
    pool.columns = columns

    normalized = canonicalize_pool_frame(pool, policy=policy, sanitize_pool=False)

    assert normalized["mentor_id"].tolist() == ["EMP-001", "EMP-002"]
    assert normalized["کد کارمندی پشتیبان"].tolist() == ["EMP-001", "EMP-002"]


def test_allocate_batch_respects_center_priority_ordering() -> None:
    policy = load_policy()
    students = pd.DataFrame(
        [
            {
                "student_id": "S-A",
                "کدرشته": 27,
                "گروه_آزمایشی": 27,
                "جنسیت": 1,
                "دانش_آموز_فارغ": 0,
                "مرکز_گلستان_صدرا": 0,
                "مالی_حکمت_بنیاد": 0,
                "کد_مدرسه": 3581,
            },
            {
                "student_id": "S-B",
                "کدرشته": 27,
                "گروه_آزمایشی": 27,
                "جنسیت": 1,
                "دانش_آموز_فارغ": 0,
                "مرکز_گلستان_صدرا": 2,
                "مالی_حکمت_بنیاد": 0,
                "کد_مدرسه": 3581,
            },
            {
                "student_id": "S-C",
                "کدرشته": 27,
                "گروه_آزمایشی": 27,
                "جنسیت": 1,
                "دانش_آموز_فارغ": 0,
                "مرکز_گلستان_صدرا": 1,
                "مالی_حکمت_بنیاد": 0,
                "کد_مدرسه": 3581,
            },
        ]
    )
    pool = pd.DataFrame(
        {
            "پشتیبان": ["M1", "M2", "M3"],
            "کد کارمندی پشتیبان": ["EMP-1", "EMP-2", "EMP-3"],
            "کدرشته": [27, 27, 27],
            "کدرشته | group_code": [27, 27, 27],
            "گروه آزمایشی": [27, 27, 27],
            "گروه آزمایشی | exam_group": [27, 27, 27],
            "جنسیت": [1, 1, 1],
            "جنسیت | gender": [1, 1, 1],
            "دانش آموز فارغ": [0, 0, 0],
            "دانش آموز فارغ | graduation_status": [0, 0, 0],
            "مرکز گلستان صدرا": [0, 2, 1],
            "مرکز گلستان صدرا | center": [0, 2, 1],
            "مالی حکمت بنیاد": [0, 0, 0],
            "مالی حکمت بنیاد | finance": [0, 0, 0],
            "کد مدرسه": [3581, 3581, 3581],
            "کد مدرسه | school_code": [3581, 3581, 3581],
            "remaining_capacity": [5, 5, 5],
            "occupancy_ratio": [0.5, 0.4, 0.3],
            "allocations_new": [0, 0, 0],
        }
    )

    _, _, logs, _ = allocate_batch(
        students,
        pool,
        policy=policy,
        center_priority=[2, 1, 0],
    )

    assert logs["student_id"].tolist() == ["S-B", "S-C", "S-A"]


def test_allocate_batch_aligns_group_crosswalk_for_students_and_pool() -> None:
    policy = load_policy()
    crosswalk = pd.DataFrame(
        {
            "گروه آزمایشی": ["علوم پایه", "علوم پایه"],
            "کد گروه": [5, 5],
            "کدرشته خام": [7, 8],
            "مقطع تحصیلی": ["دهم", "یازدهم"],
        }
    )
    group_crosswalk = build_student_group_crosswalk(crosswalk)
    students = pd.DataFrame(
        {
            "student_id": ["S-X", "S-Y"],
            "کدرشته": [7, 8],
            "گروه آزمایشی": ["علوم پایه", "علوم پایه"],
            "جنسیت": [1, 1],
            "دانش آموز فارغ": [0, 0],
            "مرکز گلستان صدرا": [1, 1],
            "مالی حکمت بنیاد": [0, 0],
            "کد مدرسه": [3581, 3581],
        }
    )
    pool = pd.DataFrame(
        {
            "پشتیبان": ["M1", "M2"],
            "کد کارمندی پشتیبان": ["EMP-10", "EMP-20"],
            "کدرشته": [5, 5],
            "کدرشته | group_code": [5, 5],
            "گروه آزمایشی": ["علوم پایه", "علوم پایه"],
            "گروه آزمایشی | exam_group": ["علوم پایه", "علوم پایه"],
            "جنسیت": [1, 1],
            "جنسیت | gender": [1, 1],
            "دانش آموز فارغ": [0, 0],
            "دانش آموز فارغ | graduation_status": [0, 0],
            "مرکز گلستان صدرا": [1, 1],
            "مرکز گلستان صدرا | center": [1, 1],
            "مالی حکمت بنیاد": [0, 0],
            "مالی حکمت بنیاد | finance": [0, 0],
            "کد مدرسه": [3581, 3581],
            "کد مدرسه | school_code": [3581, 3581],
            "remaining_capacity": [1, 1],
            "occupancy_ratio": [0.1, 0.2],
            "allocations_new": [0, 0],
        }
    )

    students_norm = canonicalize_students_frame(
        students, policy=policy, group_code_crosswalk=group_crosswalk
    )
    pool_norm = canonicalize_pool_frame(pool, policy=policy, sanitize_pool=False)

    _, _, logs, _ = allocate_batch(students_norm, pool_norm, policy=policy)

    log_records = logs.to_dict("records") if hasattr(logs, "to_dict") else list(logs)
    for log in log_records:
        join_keys = log["join_keys"]
        assert join_keys["کدرشته"] == 5
        assert log["candidate_count"] == 2


def test_allocate_batch_filters_by_center_manager() -> None:
    policy = load_policy()
    students = _single_student(student_id="S-1")
    pool = pd.DataFrame(
        {
            "پشتیبان": ["هدف", "دیگر"],
            "کد کارمندی پشتیبان": ["EMP-10", "EMP-20"],
            "کدرشته": [27, 27],
            "کدرشته | group_code": [27, 27],
            "گروه آزمایشی": [27, 27],
            "گروه آزمایشی | exam_group": [27, 27],
            "جنسیت": [1, 1],
            "جنسیت | gender": [1, 1],
            "دانش آموز فارغ": [0, 0],
            "دانش آموز فارغ | graduation_status": [0, 0],
            "مرکز گلستان صدرا": [1, 1],
            "مرکز گلستان صدرا | center": [1, 1],
            "مالی حکمت بنیاد": [0, 0],
            "مالی حکمت بنیاد | finance": [0, 0],
            "کد مدرسه": [3581, 3581],
            "کد مدرسه | school_code": [3581, 3581],
            "remaining_capacity": [1, 1],
            "occupancy_ratio": [0.9, 0.1],
            "allocations_new": [5, 0],
            "مدیر": ["شهدخت کشاورز", "مدیر دیگر"],
        }
    )

    allocations, _, logs, _ = allocate_batch(
        students,
        pool,
        policy=policy,
        center_manager_map={1: ("شهدخت کشاورز",)},
    )

    assert allocations["mentor_id"].iloc[0] == "EMP-10"
    assert logs.loc[0, "mentor_selected"].strip() == "هدف"


def test_school_students_processed_first() -> None:
    policy = load_policy()
    school_student = _single_student(student_id="S-school")
    central_student = _single_student(student_id="S-center", **{"کد_مدرسه": 0})
    mixed = pd.concat([central_student, school_student, central_student], ignore_index=True)
    pool = pd.DataFrame(
        {
            "پشتیبان": ["A", "B"],
            "کد کارمندی پشتیبان": ["EMP-1", "EMP-2"],
            "کدرشته": [27, 27],
            "کدرشته | group_code": [27, 27],
            "گروه آزمایشی": [27, 27],
            "جنسیت": [1, 1],
            "دانش آموز فارغ": [0, 0],
            "مرکز گلستان صدرا": [1, 0],
            "مرکز گلستان صدرا | center": [1, 0],
            "مالی حکمت بنیاد": [0, 0],
            "کد مدرسه": [3581, 0],
            "کد مدرسه | school_code": [3581, 0],
            "remaining_capacity": [5, 5],
            "occupancy_ratio": [0.2, 0.1],
            "allocations_new": [0, 0],
        }
    )

    _, _, logs, _ = allocate_batch(mixed, pool, policy=policy)

    assert logs["student_id"].tolist()[0] == "S-school"


def test_invalid_center_value_adds_alert(_base_pool: pd.DataFrame) -> None:
    policy = load_policy()
    students = _single_student(student_id="S-1", مرکز_گلستان_صدرا="نامعتبر")

    _, _, logs, _ = allocate_batch(students, _base_pool, policy=policy)

    alerts = logs.iloc[0]["alerts"]
    assert isinstance(alerts, list)
    assert any(alert.get("code") == "INVALID_CENTER" for alert in alerts)


def test_allocate_with_invalid_center_values(_base_pool: pd.DataFrame) -> None:
    students = pd.concat(
        [
            _single_student(student_id="S-1", مرکز_گلستان_صدرا=1),
            _single_student(student_id="S-2", مرکز_گلستان_صدرا="invalid"),
            _single_student(student_id="S-3", مرکز_گلستان_صدرا=None),
            _single_student(student_id="S-4", مرکز_گلستان_صدرا=""),
        ],
        ignore_index=True,
    )

    allocations, _, logs, _ = allocate_batch(students, _base_pool)

    assert isinstance(allocations, pd.DataFrame)
    invalid_column = logs.get("invalid_center_alerts")
    assert invalid_column is not None
    assert any(isinstance(value, list) and value for value in invalid_column)


def test_missing_manager_validation_strict() -> None:
    policy = load_policy()
    strict_policy = replace(
        policy,
        center_management=replace(
            policy.center_management,
            strict_manager_validation=True,
        ),
    )
    students = _single_student(student_id="S-1")
    pool = pd.DataFrame(
        {
            "پشتیبان": ["X"],
            "کد کارمندی پشتیبان": ["EMP-10"],
            "کدرشته": [27],
            "کدرشته | group_code": [27],
            "گروه آزمایشی": [27],
            "جنسیت": [1],
            "دانش آموز فارغ": [0],
            "مرکز گلستان صدرا": [1],
            "مرکز گلستان صدرا | center": [1],
            "مالی حکمت بنیاد": [0],
            "کد مدرسه": [3581],
            "کد مدرسه | school_code": [3581],
            "remaining_capacity": [2],
            "allocations_new": [0],
            "occupancy_ratio": [0.0],
            "مدیر": ["دیگری"],
        }
    )

    with pytest.raises(ValueError):
        allocate_batch(
            students,
            pool,
            policy=strict_policy,
            center_manager_map={1: ("ناموجود",)},
        )


def test_allocate_with_missing_school_column(_base_pool: pd.DataFrame) -> None:
    policy = load_policy()
    policy_missing = replace(
        policy,
        center_management=replace(
            policy.center_management,
            school_student_column="column_that_doesnt_exist",
        ),
    )
    students = _single_student(student_id="S-missing")

    allocations, _, _, _ = allocate_batch(students, _base_pool, policy=policy_missing)

    assert isinstance(allocations, pd.DataFrame)


def test_allocate_with_empty_manager_pool(_base_pool: pd.DataFrame) -> None:
    pool = _base_pool.copy()
    pool["manager_name"] = pd.NA
    students = _single_student(student_id="S-empty")

    allocations, _, _, _ = allocate_batch(
        students,
        pool,
        center_manager_map={1: ("مدیر خاص",)},
    )

    assert isinstance(allocations, pd.DataFrame)


def test_canonicalize_pool_frame_reports_join_key_duplicates(
    _base_pool: pd.DataFrame,
) -> None:
    policy = load_policy()

    duplicated = pd.concat([_base_pool, _base_pool.iloc[[0]].copy()], ignore_index=True)

    normalized = canonicalize_pool_frame(duplicated, policy=policy, sanitize_pool=False)
    duplicate_report = normalized.attrs[POOL_JOIN_KEY_DUPLICATES_ATTR]
    stats = normalized.attrs["pool_canonicalization_stats"]
    summary = normalized.attrs[POOL_DUPLICATE_SUMMARY_ATTR]

    assert not duplicate_report.empty
    assert duplicate_report["کد کارمندی پشتیبان"].tolist() == ["EMP-001", "EMP-001"]
    assert summary["duplicate_scope"] == "per_mentor"
    assert normalized.attrs.get("pool_duplicate_scope") == "per_mentor"
    assert stats.join_key_duplicates == len(duplicate_report)
    assert summary["total"] == len(duplicate_report)
    assert isinstance(summary["sample"], list)


def test_canonicalize_pool_frame_allows_distinct_mentors_same_join_keys(
    _base_pool: pd.DataFrame,
) -> None:
    policy = load_policy()

    normalized = canonicalize_pool_frame(_base_pool, policy=policy, sanitize_pool=False)
    duplicate_report = normalized.attrs[POOL_JOIN_KEY_DUPLICATES_ATTR]

    assert duplicate_report.empty
    assert normalized.attrs["pool_duplicate_scope"] == "per_mentor"
    summary = normalized.attrs[POOL_DUPLICATE_SUMMARY_ATTR]
    assert summary["duplicate_scope"] == "per_mentor"
    assert summary["total"] == 0


def test_canonicalize_pool_frame_reports_key_level_duplicates_when_requested(
    _base_pool: pd.DataFrame,
) -> None:
    policy = load_policy()

    normalized = canonicalize_pool_frame(
        _base_pool, policy=policy, sanitize_pool=False, include_distinct_mentor_duplicates=True
    )
    duplicate_report = normalized.attrs[POOL_JOIN_KEY_DUPLICATES_ATTR]
    summary = normalized.attrs[POOL_DUPLICATE_SUMMARY_ATTR]

    assert not duplicate_report.empty
    assert duplicate_report["کد کارمندی پشتیبان"].tolist() == ["EMP-001", "EMP-002"]
    assert duplicate_report["duplicate_group_size"].dropna().unique().tolist() == [2]
    assert duplicate_report.attrs.get("duplicate_scope") == "per_key"
    assert normalized.attrs.get("pool_duplicate_scope") == "per_key"
    assert summary["duplicate_scope"] == "per_key"


def test_build_join_key_duplicate_report_counts_only_repeated_mentor_rows(
    _base_pool: pd.DataFrame,
) -> None:
    join_keys = [
        "کدرشته",
        "جنسیت",
        "دانش آموز فارغ",
        "مرکز گلستان صدرا",
        "مالی حکمت بنیاد",
        "کد مدرسه",
    ]
    mentor_column = "کد کارمندی پشتیبان"

    pool_subset = _base_pool[[*join_keys, mentor_column]].copy()
    repeated = pd.concat(
        [
            pool_subset,
            pool_subset.iloc[[0]].copy(),
            pool_subset.iloc[[1]].copy(),
            pool_subset.iloc[[1]].copy(),
        ],
        ignore_index=True,
    )
    repeated.loc[len(repeated)] = {
        "کدرشته": 27,
        "جنسیت": 1,
        "دانش آموز فارغ": 0,
        "مرکز گلستان صدرا": 1,
        "مالی حکمت بنیاد": 0,
        "کد مدرسه": 3581,
        "کد کارمندی پشتیبان": "EMP-999",
    }

    report = _build_join_key_duplicate_report(
        repeated, join_keys, mentor_column, include_distinct_mentors=False
    )

    assert not report.empty
    assert set(report[mentor_column].unique()) == {"EMP-001", "EMP-002"}
    assert sorted(report["duplicate_group_size"].dropna().unique().tolist()) == [2, 3]
    assert set(report.columns) == set(join_keys + [mentor_column, "duplicate_group_size"])


def test_build_join_key_duplicate_report_handles_multiple_groups_per_mode() -> None:
    policy = load_policy()
    join_keys = policy.join_keys
    mentor_column = "کد کارمندی پشتیبان"
    rows = [
        {
            "کدرشته": 11,
            "جنسیت": 0,
            "دانش آموز فارغ": 0,
            "مرکز گلستان صدرا": 1,
            "مالی حکمت بنیاد": 0,
            "کد مدرسه": 4001,
            mentor_column: "EMP-A",
        },
        {
            "کدرشته": 11,
            "جنسیت": 0,
            "دانش آموز فارغ": 0,
            "مرکز گلستان صدرا": 1,
            "مالی حکمت بنیاد": 0,
            "کد مدرسه": 4001,
            mentor_column: "EMP-B",
        },
        {
            "کدرشته": 9,
            "جنسیت": 1,
            "دانش آموز فارغ": 0,
            "مرکز گلستان صدرا": 2,
            "مالی حکمت بنیاد": 0,
            "کد مدرسه": 5001,
            mentor_column: "EMP-C",
        },
        {
            "کدرشته": 9,
            "جنسیت": 1,
            "دانش آموز فارغ": 0,
            "مرکز گلستان صدرا": 2,
            "مالی حکمت بنیاد": 0,
            "کد مدرسه": 5001,
            mentor_column: "EMP-C",
        },
    ]
    frame = pd.DataFrame(rows)

    per_mentor = _build_join_key_duplicate_report(frame, join_keys, mentor_column)
    per_key = _build_join_key_duplicate_report(
        frame, join_keys, mentor_column, include_distinct_mentors=True
    )

    assert per_mentor[mentor_column].tolist() == ["EMP-C", "EMP-C"]
    assert per_mentor["duplicate_group_size"].dropna().unique().tolist() == [2]

    assert per_key[mentor_column].tolist() == ["EMP-A", "EMP-B", "EMP-C", "EMP-C"]
    group_sizes = per_key.groupby(join_keys, sort=False)["duplicate_group_size"].first()
    assert group_sizes.to_dict() == {
        tuple(rows[0][k] for k in join_keys): 2,
        tuple(rows[2][k] for k in join_keys): 2,
    }


def test_build_join_key_duplicate_report_missing_columns_returns_empty() -> None:
    join_keys = [
        "کدرشته",
        "جنسیت",
        "دانش آموز فارغ",
        "مرکز گلستان صدرا",
        "مالی حکمت بنیاد",
        "کد مدرسه",
    ]
    mentor_column = "کد کارمندی پشتیبان"
    frame = pd.DataFrame(
        {
            "کدرشته": [27],
            "جنسیت": [1],
            "دانش آموز فارغ": [0],
            "مرکز گلستان صدرا": [1],
            "مالی حکمت بنیاد": [0],
            mentor_column: ["EMP-001"],
        }
    )

    report = _build_join_key_duplicate_report(frame, join_keys, mentor_column)
    with_distinct = _build_join_key_duplicate_report(
        frame, join_keys, mentor_column, include_distinct_mentors=True
    )

    for candidate in (report, with_distinct):
        assert candidate.empty
        assert candidate.columns.tolist() == join_keys + [mentor_column, "duplicate_group_size"]


def test_sanitize_pool_records_virtual_and_capacity_stats() -> None:
    policy = load_policy()
    raw = pd.DataFrame(
        {
            "mentor_name": ["فراگیر آزمون", "علی"],
            "alias": [7505, 102],
            "remaining_capacity": ["5", "X"],
        }
    )

    sanitized = sanitize_pool_for_allocation(raw, policy=policy)
    stats = sanitized.attrs["pool_canonicalization_stats"]

    assert stats.virtual_filtered == 1
    assert stats.capacity_coerced == 1


def test_canonicalize_pool_frame_records_mentor_id_autofill(_base_pool: pd.DataFrame) -> None:
    policy = load_policy()
    pool = _base_pool.drop(columns=["mentor_id"], errors="ignore")

    normalized = canonicalize_pool_frame(pool, policy=policy, sanitize_pool=False)
    stats = normalized.attrs["pool_canonicalization_stats"]

    assert stats.mentor_id_autofill == len(pool)


def test_canonicalize_pool_frame_uses_alias_map_for_missing_ids(
    _base_pool: pd.DataFrame,
) -> None:
    policy = load_policy()
    pool = _base_pool.copy()
    pool.loc[1, "کد کارمندی پشتیبان"] = ""
    pool["alias"] = ["0012345678", "0012345678"]

    normalized = canonicalize_pool_frame(pool, policy=policy, sanitize_pool=False)
    stats = normalized.attrs["pool_canonicalization_stats"]

    assert normalized.loc[1, "mentor_id"] == "EMP-001"
    assert stats.alias_autofill == 1
    assert stats.alias_unmatched == 0


def test_canonicalize_pool_frame_tracks_alias_unmatched(_base_pool: pd.DataFrame) -> None:
    policy = load_policy()
    pool = _base_pool.copy()
    pool.loc[1, "کد کارمندی پشتیبان"] = ""
    pool["alias"] = ["0012345678", "009876543"]

    normalized = canonicalize_pool_frame(pool, policy=policy, sanitize_pool=False)
    stats = normalized.attrs["pool_canonicalization_stats"]

    assert stats.alias_unmatched == 1


def test_allocate_student_dict_missing_school_field_skips_filter(
    _base_pool: pd.DataFrame,
) -> None:
    student_row = _single_student().iloc[0].to_dict()
    student_row.pop("کد_مدرسه")
    student_row["school_code_norm"] = None

    result = allocate_student(student_row, _base_pool)

    assert result.log["allocation_status"] == "success"
    assert result.log["error_type"] is None
    assert result.log["join_keys"]["کد_مدرسه"] == 0
    school_trace = next(stage for stage in result.trace if stage["stage"] == "school")
    assert school_trace["total_after"] == school_trace["total_before"]


@pytest.mark.parametrize("raw_code", ["35-81", "35/81", "35\\81", "۳۵-۸۱", "35–81"])
def test_allocate_student_sanitizes_school_code_separators(
    raw_code: str, _base_pool: pd.DataFrame
) -> None:
    student_row = _single_student(کد_مدرسه=raw_code).iloc[0].to_dict()

    result = allocate_student(student_row, _base_pool)

    assert result.log["allocation_status"] == "success"
    assert result.log["error_type"] is None
    assert result.log["join_keys"]["کد_مدرسه"] == 3581


def test_allocate_student_with_string_join_values_matches(_base_pool: pd.DataFrame) -> None:
    student_row = _single_student().iloc[0].to_dict()
    for key in ("کدرشته", "جنسیت", "دانش_آموز_فارغ", "مرکز_گلستان_صدرا", "مالی_حکمت_بنیاد"):
        if key in student_row:
            student_row[key] = str(student_row[key])
    student_row["کد_مدرسه"] = "3581"


def test_allocate_student_center_zero_skips_filter(_base_pool: pd.DataFrame) -> None:
    student_row = _single_student(مرکز_گلستان_صدرا=0).iloc[0].to_dict()

    result = allocate_student(student_row, _base_pool)

    assert result.log["allocation_status"] == "failed"
    assert result.log["error_type"] == "ELIGIBILITY_NO_MATCH"
    assert result.log["candidate_count"] == 0
    assert result.log["rule_reason_code"] == "CENTER_MISMATCH"


def test_allocate_batch_skips_canonicalization_when_frames_prepared(
    _base_pool: pd.DataFrame, monkeypatch: pytest.MonkeyPatch
) -> None:
    policy = load_policy()
    students = _single_student()
    calls = {"students": 0, "pool": 0}

    def _spy_students(df: pd.DataFrame, policy: PolicyConfig) -> pd.DataFrame:
        calls["students"] += 1
        return canonicalize_students_frame(df, policy=policy)

    def _spy_pool(df: pd.DataFrame, policy: PolicyConfig) -> pd.DataFrame:
        calls["pool"] += 1
        return canonicalize_pool_frame(
            df,
            policy=policy,
            sanitize_pool=False,
            pool_source="inspactor",
        )

    monkeypatch.setattr("app.core.allocate_students._normalize_students", _spy_students)
    monkeypatch.setattr("app.core.allocate_students._normalize_pool", _spy_pool)

    allocate_batch(students.copy(deep=True), _base_pool.copy(deep=True), policy=policy)
    assert calls == {"students": 1, "pool": 1}

    students_canon, pool_canon = canonicalize_allocation_frames(
        students.copy(deep=True),
        _base_pool.copy(deep=True),
        policy=policy,
        sanitize_pool=False,
        pool_source="inspactor",
    )

    allocate_batch(
        students_canon,
        pool_canon,
        policy=policy,
        frames_already_canonical=True,
    )

    assert calls == {"students": 1, "pool": 1}


def test_allocate_batch_no_match_sets_error(_base_pool: pd.DataFrame) -> None:
    students = _single_student(**{"کدرشته": 33})

    allocations, updated_pool, logs, _ = allocate_batch(students, _base_pool)

    assert allocations.empty
    assert list(allocations.columns) == [
        "student_id",
        "student_national_code",
        "mentor",
        "mentor_id",
        "mentor_alias_code",
    ]
    pd.testing.assert_frame_equal(updated_pool[_base_pool.columns], _base_pool, check_dtype=False)
    assert "school_code" in updated_pool.columns
    assert logs.iloc[0]["error_type"] == "ELIGIBILITY_NO_MATCH"
    assert logs.iloc[0]["detailed_reason"] == "No candidates matched join keys"


def test_normalize_pool_appends_pipe_alias_columns() -> None:
    policy = load_policy()
    pool = pd.DataFrame(
        {
            "پشتیبان": ["زهرا"],
            "کد کارمندی پشتیبان": ["EMP-001"],
            "کدرشته": [27],
            "جنسیت": [1],
            "دانش آموز فارغ": [0],
            "مرکز گلستان صدرا": [1],
            "مالی حکمت بنیاد": [0],
            "کد مدرسه": [3581],
            "remaining_capacity": [2],
            "allocations_new": [0],
            "occupancy_ratio": [0.0],
        }
    )

    normalized = _normalize_pool(pool, policy)

    expected_pairs = [
        ("کدرشته", "کدرشته | group_code"),
        ("جنسیت", "جنسیت | gender"),
        ("دانش آموز فارغ", "دانش آموز فارغ | graduation_status"),
        ("مرکز گلستان صدرا", "مرکز گلستان صدرا | center"),
        ("مالی حکمت بنیاد", "مالی حکمت بنیاد | finance"),
        ("کد مدرسه", "کد مدرسه | school_code"),
    ]

    for fa_name, bilingual in expected_pairs:
        assert bilingual in normalized.columns
        pd.testing.assert_series_equal(
            normalized[fa_name],
            normalized[bilingual],
            check_dtype=False,
            check_names=False,
        )


def test_allocate_batch_capacity_full_sets_error(_base_pool: pd.DataFrame) -> None:
    students = _single_student()
    pool = _base_pool.assign(remaining_capacity=[0, 0])

    allocations, updated_pool, logs, _ = allocate_batch(students, pool)

    assert allocations.empty
    assert (updated_pool["remaining_capacity"] == 0).all()
    assert logs.iloc[0]["error_type"] == "CAPACITY_FULL"
    assert logs.iloc[0]["candidate_count"] == 2
    assert logs.iloc[0]["detailed_reason"] == "No capacity among matched candidates"


def test_allocate_batch_logs_include_alias_stats(_base_pool: pd.DataFrame) -> None:
    students = _single_student()
    pool = _base_pool.copy()
    pool.loc[1, "کد کارمندی پشتیبان"] = ""
    pool["alias"] = ["0012345678", "0012345678"]

    _, _, logs, _ = allocate_batch(students, pool)

    assert logs["alias_autofill"].unique().tolist() == [1]
    assert logs["alias_unmatched"].unique().tolist() == [0]


def test_canonicalize_allocation_frames_accepts_english_join_keys() -> None:
    policy = load_policy()
    students = _single_student()
    pool = pd.DataFrame(
        {
            "mentor_name": ["زهرا"],
            "mentor_id": ["EMP-01"],
            "group_code": [27],
            "gender": [1],
            "graduation_status": [0],
            "center": [1],
            "finance": [0],
            "school_code": [3581],
            "remaining_capacity": [2],
        }
    )

    students_canon, pool_canon = canonicalize_allocation_frames(
        students,
        pool,
        policy=policy,
        sanitize_pool=True,
        pool_source="inspactor",
    )

    expected_values = {
        "کدرشته": 27,
        "جنسیت": 1,
        "دانش آموز فارغ": 0,
        "مرکز گلستان صدرا": 1,
        "مالی حکمت بنیاد": 0,
        "کد مدرسه": 3581,
    }

    for join_key, expected in expected_values.items():
        assert join_key in pool_canon.columns
        assert pool_canon[join_key].tolist() == [expected]
    assert students_canon is not None


def test_allocate_student_handles_empty_ranking(
    monkeypatch: pytest.MonkeyPatch, _base_pool: pd.DataFrame
) -> None:
    student = _single_student().iloc[0].to_dict()

    def _empty_ranked(df: pd.DataFrame, **_: object) -> pd.DataFrame:
        empty = df.iloc[0:0].copy()
        empty.attrs["fairness_reason"] = None
        return empty

    monkeypatch.setattr("app.core.allocate_students.apply_ranking_policy", _empty_ranked)

    result = allocate_student(student, _base_pool)

    assert result.mentor_row is None
    assert result.log["error_type"] == "INTERNAL_ERROR"
    assert "Ranking policy returned no candidates" in str(result.log["detailed_reason"])


def test_allocate_student_handles_canonicalization_empty(
    monkeypatch: pytest.MonkeyPatch, _base_pool: pd.DataFrame
) -> None:
    student = _single_student().iloc[0].to_dict()

    import app.core.allocate_students as module

    original_canonicalize = module.canonicalize_headers
    original_apply_ranking = module.apply_ranking_policy

    def _mark_ranked(*args: object, **kwargs: object) -> pd.DataFrame:
        ranked = original_apply_ranking(*args, **kwargs)
        ranked["__force_canon_empty__"] = 1
        return ranked

    def _empty_ranked(df: pd.DataFrame, **kwargs: object) -> pd.DataFrame:
        result = original_canonicalize(df, **kwargs)
        if isinstance(df, pd.DataFrame) and "__force_canon_empty__" in df.columns:
            return result.iloc[0:0]
        return result

    monkeypatch.setattr(module, "apply_ranking_policy", _mark_ranked)
    monkeypatch.setattr(module, "canonicalize_headers", _empty_ranked)

    result = allocate_student(student, _base_pool)

    assert result.mentor_row is None
    assert result.log["error_type"] == "INTERNAL_ERROR"
    assert "Canonicalization returned empty ranked view" in str(result.log["detailed_reason"])


def test_allocate_batch_progress_reports_start_and_end(_base_pool: pd.DataFrame) -> None:
    students = pd.concat(
        [_single_student(), _single_student(student_id="STD-002")], ignore_index=True
    )
    progress_calls: list[tuple[int, str]] = []

    def _progress(pct: int, msg: str) -> None:
        progress_calls.append((pct, msg))

    allocate_batch(students, _base_pool, progress=_progress)

    assert progress_calls[0][0] == 0
    assert progress_calls[0][1] == "start"
    assert any(pct == 100 for pct, _ in progress_calls)
    assert progress_calls[-1][1] == "done"


def test_allocate_student_records_fairness_reason_code(_base_pool: pd.DataFrame) -> None:
    policy = load_policy()
    policy = replace(policy, fairness_strategy="deterministic_jitter")
    pool = _base_pool.assign(
        occupancy_ratio=[0.0, 0.0],
        allocations_new=[0, 0],
        **{"کد کارمندی پشتیبان": ["EMP1", "EMP01"]},
        counter=["543570002", "543570001"],
    )

    student_row = _single_student().iloc[0].to_dict()

    result = allocate_student(student_row, pool, policy=policy)

    assert result.log["fairness_reason_code"] == "FAIRNESS_ORDER"
    fairness_text = result.log.get("fairness_reason_text") or ""
    assert "[FAIRNESS_ORDER]" in fairness_text


def test_allocate_student_records_mentor_state_delta(_base_pool: pd.DataFrame) -> None:
    student_row = _single_student().iloc[0].to_dict()
    state = {
        "EMP-001": {"initial": 2, "remaining": 2, "alloc_new": 0, "occupancy_ratio": 0.0},
        "EMP-002": {"initial": 2, "remaining": 2, "alloc_new": 0, "occupancy_ratio": 0.0},
    }

    result = allocate_student(student_row, _base_pool, state=state)

    delta = result.log.get("mentor_state_delta")
    assert delta is not None
    assert delta["before"]["remaining"] == 2
    assert delta["after"]["remaining"] == 1
    assert delta["diff"]["remaining"] == -1
    assert delta["diff"]["alloc_new"] == 1
    assert delta["after"]["occupancy_ratio"] == 0.0


def test_allocate_student_underflow_embeds_snapshot_details(
    _base_pool: pd.DataFrame,
) -> None:
    student_row = _single_student().iloc[0].to_dict()
    pool = _base_pool.iloc[[0]].copy()
    state = {
        "EMP-001": {"initial": 1, "remaining": 0, "alloc_new": 1, "occupancy_ratio": 1.0},
        "EMP-002": {"initial": 2, "remaining": 2, "alloc_new": 0, "occupancy_ratio": 0.0},
    }

    result = allocate_student(student_row, pool, state=state)

    assert result.log["error_type"] == "CAPACITY_UNDERFLOW"
    reason = result.log.get("detailed_reason") or ""
    assert "student=STD-001" in reason
    assert "mentor=EMP-001" in reason
    assert "mentor snapshot" in reason
    delta = result.log.get("mentor_state_delta")
    assert delta is not None
    assert delta["before"]["remaining"] == 0
    assert delta["diff"]["alloc_new"] == 0


def test_allocate_batch_handles_underflow_without_unboundlocal(
    _base_pool: pd.DataFrame,
) -> None:
    students = _single_student()
    pool = _base_pool.iloc[[0]].copy()
    pool["remaining_capacity"] = [0]

    _, _, logs, _ = allocate_batch(students, pool)

    assert len(logs) == 1
    first_log = logs.iloc[0]
    assert first_log["error_type"] in {"CAPACITY_FULL", "CAPACITY_UNDERFLOW"}
    assert first_log.get("mentor_id") is None


def test_join_key_values_validates_length() -> None:
    with pytest.raises(ValueError):
        JoinKeyValues({"a": 1, "b": 2, "c": 3, "d": 4, "e": 5})


def test_join_key_values_rejects_non_int() -> None:
    payload = {key: index for index, key in enumerate(CANONICAL_JOIN_KEYS, start=1)}
    payload[CANONICAL_JOIN_KEYS[2]] = "oops"
    with pytest.raises(TypeError):
        JoinKeyValues(payload)


def test_allocate_batch_join_keys_are_typed(_base_pool: pd.DataFrame) -> None:
    students = _single_student()

    _, _, logs, _ = allocate_batch(students, _base_pool)

    join_values = logs.iloc[0]["join_keys"]
    assert isinstance(join_values, JoinKeyValues)
    assert list(join_values.keys()) == list(CANONICAL_JOIN_KEYS)


def test_allocate_batch_missing_school_code_defaults_to_zero(
    _base_pool: pd.DataFrame,
) -> None:
    students = _single_student(**{"کد_مدرسه": None})

    allocations, updated_pool, logs, _ = allocate_batch(students, _base_pool)

    assert len(allocations) == 1
    assert allocations.iloc[0]["mentor_id"] == "EMP-001"
    assert int(updated_pool.loc[0, "remaining_capacity"]) == 1
    record = logs.iloc[0]
    assert record["allocation_status"] == "success"
    assert record["error_type"] is None
    join_values = record["join_keys"]
    assert isinstance(join_values, JoinKeyValues)
    assert join_values["کد_مدرسه"] == 0


def test_allocate_batch_missing_school_code_requires_data_when_disabled(
    _base_pool: pd.DataFrame,
) -> None:
    payload = json.loads(Path("config/policy.json").read_text(encoding="utf-8"))
    payload["school_code_empty_as_zero"] = False
    policy = parse_policy_dict(payload)

    students = _single_student(**{"کد_مدرسه": None})

    allocations, updated_pool, logs, _ = allocate_batch(students, _base_pool, policy=policy)

    assert allocations.empty
    pd.testing.assert_frame_equal(updated_pool[_base_pool.columns], _base_pool, check_dtype=False)
    record = logs.iloc[0]
    assert record["error_type"] == "DATA_MISSING"
    assert "کد مدرسه" in str(record["detailed_reason"])
    join_values = record["join_keys"]
    assert isinstance(join_values, JoinKeyValues)
    assert join_values["کد_مدرسه"] == -1


def test_allocate_batch_logs_capacity_transition(_base_pool: pd.DataFrame) -> None:
    students = _single_student()

    _, updated_pool, logs, _ = allocate_batch(students, _base_pool)

    assert logs.iloc[0]["capacity_before"] == 2
    assert logs.iloc[0]["capacity_after"] == 1
    assert int(updated_pool.loc[0, "remaining_capacity"]) == 1


def test_allocate_batch_reconciles_numeric_mentor_ids(_base_pool: pd.DataFrame) -> None:
    students = _single_student()
    pool = _base_pool.copy()
    pool["کد کارمندی پشتیبان"] = [101, 102]

    allocations, updated_pool, logs, _ = allocate_batch(students, pool)

    assert logs.iloc[0]["error_type"] is None
    assert allocations.iloc[0]["mentor_id"] == "101"
    assert int(updated_pool.loc[0, "remaining_capacity"]) == 1


def test_allocate_batch_handles_missing_state(
    monkeypatch: pytest.MonkeyPatch, _base_pool: pd.DataFrame
) -> None:
    students = _single_student()

    def _raise_missing_state(_: object, __: object) -> tuple[int, int, float]:
        raise KeyError("Mentor 'EMP-001' missing from state")

    monkeypatch.setattr(
        "app.core.allocate_students.consume_capacity",
        _raise_missing_state,
        raising=True,
    )

    allocations, updated_pool, logs, _ = allocate_batch(students, _base_pool)

    assert allocations.empty
    pd.testing.assert_frame_equal(updated_pool[_base_pool.columns], _base_pool, check_dtype=False)
    record = logs.iloc[0]
    assert record["allocation_status"] == "failed"
    assert record["error_type"] == "INTERNAL_ERROR"
    assert "missing" in str(record["detailed_reason"]).lower()


def test_allocate_batch_missing_capacity_column_fails_contract(
    monkeypatch: pytest.MonkeyPatch, _base_pool: pd.DataFrame
) -> None:
    students = _single_student()
    pool = _base_pool.drop(columns=["remaining_capacity"])

    def _identity_pool(df: pd.DataFrame, _: object) -> pd.DataFrame:
        return df

    monkeypatch.setattr(
        "app.core.allocate_students._normalize_pool",
        _identity_pool,
        raising=True,
    )

    with pytest.raises(ValueError, match="Canonical pool frame missing columns"):
        allocate_batch(students, pool)


def test_allocate_batch_cli_wraps_contract_error(_base_pool: pd.DataFrame) -> None:
    policy = load_policy()
    students_norm, pool_norm = canonicalize_allocation_frames(
        _single_student(),
        _base_pool,
        policy=policy,
    )
    pool_missing = pool_norm.drop(columns=["remaining_capacity"])

    with pytest.raises(ValueError, match="DATA_MISSING"):
        allocate_batch(
            students_norm,
            pool_missing,
            policy=policy,
            frames_already_canonical=True,
        )


def test_allocate_batch_invalid_join_value_sets_error(_base_pool: pd.DataFrame) -> None:
    students = _single_student(**{"کدرشته": ""})

    allocations, updated_pool, logs, _ = allocate_batch(students, _base_pool)

    assert allocations.empty
    pd.testing.assert_frame_equal(updated_pool[_base_pool.columns], _base_pool, check_dtype=False)
    record = logs.iloc[0]
    assert record["error_type"] == "DATA_MISSING"
    assert "کدرشته" in str(record["detailed_reason"])


def test_allocate_batch_handles_farsi_gender_tokens_for_male(_base_pool: pd.DataFrame) -> None:
    policy = load_policy()
    students = _single_student(**{"جنسیت": "پسر"})

    allocations, _, logs, _ = allocate_batch(students, _base_pool, policy=policy)

    assert not allocations.empty
    assert logs.iloc[0]["error_type"] is None


def test_allocate_batch_handles_farsi_gender_tokens_for_female(
    _base_pool: pd.DataFrame,
) -> None:
    policy = load_policy()
    pool = _base_pool.copy()
    pool.loc[0, "جنسیت"] = 0
    pool.loc[0, "جنسیت | gender"] = 0
    students = _single_student(**{"جنسیت": "دختر"})

    allocations, _, logs, _ = allocate_batch(students, pool, policy=policy)

    assert not allocations.empty
    assert logs.iloc[0]["mentor_id"] == pool.iloc[0]["کد کارمندی پشتیبان"]
    assert logs.iloc[0]["error_type"] is None


def test_allocate_batch_flags_non_int_join_key_values(_base_pool: pd.DataFrame) -> None:
    policy = load_policy()
    students = _single_student(**{"جنسیت": "نامعتبر"})

    allocations, _, logs, _ = allocate_batch(students, _base_pool, policy=policy)

    assert allocations.empty
    record = logs.iloc[0]
    assert record["error_type"] == "DATA_MISSING"
    assert "جنسیت" in str(record["detailed_reason"])


def test_policy_required_fields_enforced_from_config(
    _base_pool: pd.DataFrame,
) -> None:
    payload = json.loads(Path("config/policy.json").read_text(encoding="utf-8"))
    payload["required_student_fields"] = payload["join_keys"] + ["exam_group"]
    policy = parse_policy_dict(payload)

    students = _single_student()
    allocate_batch(students, _base_pool, policy=policy)

    missing_group = students.drop(columns=["گروه_آزمایشی"]).copy()
    normalized = canonicalize_students_frame(missing_group, policy=policy)

    exam_group_col = columns.CANON_EN_TO_FA["exam_group"]
    assert exam_group_col in normalized.columns
    assert normalized[exam_group_col].isna().all()
    allocate_batch(missing_group, _base_pool, policy=policy)


def test_separate_school_students_handles_missing_indicator_column() -> None:
    policy = load_policy()
    students = _single_student()

    school, center = _separate_school_students(students, policy)

    assert school.empty
    pd.testing.assert_frame_equal(center.reset_index(drop=True), students.reset_index(drop=True))


def test_school_students_have_priority_without_center_manager_filter() -> None:
    policy = load_policy()
    students = pd.DataFrame(
        [
            {
                "student_id": "STD-SCHOOL",
                "کدرشته": 27,
                "گروه_آزمایشی": 27,
                "جنسیت": 1,
                "دانش_آموز_فارغ": 0,
                "مرکز_گلستان_صدرا": 1,
                "مالی_حکمت_بنیاد": 0,
                "کد_مدرسه": 1111,
                "is_school_student": True,
            },
            {
                "student_id": "STD-CENTER",
                "کدرشته": 27,
                "گروه_آزمایشی": 27,
                "جنسیت": 1,
                "دانش_آموز_فارغ": 0,
                "مرکز_گلستان_صدرا": 1,
                "مالی_حکمت_بنیاد": 0,
                "کد_مدرسه": 1111,
                "is_school_student": False,
            },
        ]
    )
    pool = pd.DataFrame(
        [
            {
                "پشتیبان": "منتور آلفا",
                "کد کارمندی پشتیبان": "EMP-001",
                "کدرشته": 27,
                "گروه آزمایشی": 27,
                "جنسیت": 1,
                "دانش آموز فارغ": 0,
                "مرکز گلستان صدرا": 1,
                "مالی حکمت بنیاد": 0,
                "کد مدرسه": 1111,
                "remaining_capacity": 1,
                "allocations_new": 0,
                "occupancy_ratio": 0.0,
                "مدیر": "مدیر آلفا",
            },
            {
                "پشتیبان": "منتور بتا",
                "کد کارمندی پشتیبان": "EMP-002",
                "کدرشته": 27,
                "گروه آزمایشی": 27,
                "جنسیت": 1,
                "دانش آموز فارغ": 0,
                "مرکز گلستان صدرا": 1,
                "مالی حکمت بنیاد": 0,
                "کد مدرسه": 1111,
                "remaining_capacity": 1,
                "allocations_new": 0,
                "occupancy_ratio": 0.5,
                "مدیر": "مدیر بتا",
            },
        ]
    )

    allocations, _, _, _ = allocate_batch(
        students,
        pool,
        policy=policy,
        center_manager_map={1: ["مدیر بتا"]},
    )

    assert allocations.shape[0] == 2
    school_row = allocations.loc[allocations["student_id"] == "STD-SCHOOL"].iloc[0]
    center_row = allocations.loc[allocations["student_id"] == "STD-CENTER"].iloc[0]
    assert school_row["mentor_id"] == "EMP-001"
    assert center_row["mentor_id"] == "EMP-002"


def test_phase_rule_trace_records_school_and_center_events() -> None:
    policy = load_policy()
    students = pd.concat(
        [
            _single_student(student_id="STD-SCHOOL", is_school_student=True),
            _single_student(
                student_id="STD-CENTER",
                مرکز_گلستان_صدرا=2,
                is_school_student=False,
            ),
        ],
        ignore_index=True,
    )
    pool = pd.DataFrame(
        [
            {
                "پشتیبان": "منتور مدرسه",
                "کد کارمندی پشتیبان": "EMP-900",
                "کدرشته": 27,
                "کدرشته | group_code": 27,
                "گروه آزمایشی": 27,
                "گروه آزمایشی | exam_group": 27,
                "جنسیت": 1,
                "جنسیت | gender": 1,
                "دانش آموز فارغ": 0,
                "دانش آموز فارغ | graduation_status": 0,
                "مرکز گلستان صدرا": 1,
                "مرکز گلستان صدرا | center": 1,
                "مالی حکمت بنیاد": 0,
                "مالی حکمت بنیاد | finance": 0,
                "کد مدرسه": 1111,
                "کد مدرسه | school_code": 1111,
                "remaining_capacity": 1,
                "allocations_new": 0,
                "occupancy_ratio": 0.0,
            }
        ]
    )

    _, _, logs, _ = allocate_batch(students, pool, policy=policy)

    school_trace = logs.loc[logs["student_id"] == "STD-SCHOOL", "phase_rule_trace"].iloc[0]
    assert any(entry.get("stage") == "school_phase_start" for entry in school_trace)
    assert any(
        entry.get("reason") == ReasonCode.SCHOOL_STUDENT_PRIORITY.value for entry in school_trace
    )
    center_trace = logs.loc[logs["student_id"] == "STD-CENTER", "phase_rule_trace"].iloc[0]
    assert any(entry.get("stage") == "center_phase_start" for entry in center_trace)
    assert any(
        entry.get("reason") == ReasonCode.NO_MANAGER_FOR_CENTER.value for entry in center_trace
    )


@pytest.mark.skipif(importlib.util.find_spec("openpyxl") is None, reason="openpyxl لازم است")
def test_allocation_outputs_excel_openable(tmp_path: Path, _base_pool: pd.DataFrame) -> None:
    from openpyxl import load_workbook

    from app.infra.io_utils import write_xlsx_atomic

    students = pd.concat(
        [_single_student(), _single_student(student_id="STD-002")],
        ignore_index=True,
    )

    policy = load_policy()
    allocations, updated_pool, logs, trace = allocate_batch(students, _base_pool, policy=policy)
    reasons = build_selection_reason_rows(
        allocations,
        students,
        _base_pool,
        policy=policy,
        logs=logs,
        trace=trace,
    )
    _, reasons = write_selection_reasons_sheet(reasons, writer=None, policy=policy)

    out_path = tmp_path / "allocation_bundle.xlsx"
    write_xlsx_atomic(
        {
            "allocations": allocations,
            "pool": updated_pool,
            "logs": logs,
            "trace": trace,
            "دلایل انتخاب پشتیبان": reasons,
        },
        out_path,
    )

    workbook = load_workbook(out_path)
    assert set(workbook.sheetnames) == {
        "allocations",
        "pool",
        "logs",
        "trace",
        "دلایل انتخاب پشتیبان",
    }


@pytest.mark.skipif(
    importlib.util.find_spec("openpyxl") is None and importlib.util.find_spec("xlsxwriter") is None,
    reason="نیاز به یکی از موتورهای Excel (openpyxl/xlsxwriter)",
)
def test_cli_capacity_column_default_from_policy(tmp_path: Path) -> None:
    from app.infra import cli

    policy_path = tmp_path / "policy.json"
    payload = json.loads(Path("config/policy.json").read_text(encoding="utf-8"))
    payload["columns"]["remaining_capacity"] = "ظرفیت"
    with policy_path.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False)

    students_path = tmp_path / "students.xlsx"
    pool_path = tmp_path / "pool.xlsx"
    output_path = tmp_path / "out.xlsx"

    pd.DataFrame(
        [
            {
                "student_id": "S1",
                "national_id": "0012345678",
                "کدرشته": 27,
                "گروه_آزمایشی": 27,
                "جنسیت": 1,
                "دانش_آموز_فارغ": 0,
                "مرکز_گلستان_صدرا": 1,
                "مالی_حکمت_بنیاد": 0,
                "کد_مدرسه": 100,
            }
        ]
    ).to_excel(students_path, index=False)

    pd.DataFrame(
        [
            {
                "پشتیبان": "زهرا",
                "کد کارمندی پشتیبان": "EMP-1",
                "کدرشته": 27,
                "گروه آزمایشی": 27,
                "جنسیت": 1,
                "دانش آموز فارغ": 0,
                "مرکز گلستان صدرا": 1,
                "مالی حکمت بنیاد": 0,
                "کد مدرسه": 100,
                "ظرفیت": 1,
                "occupancy_ratio": 0.2,
                "allocations_new": 0,
            }
        ]
    ).to_excel(pool_path, index=False)

    rc = cli.main(
        [
            "allocate",
            "--students",
            str(students_path),
            "--pool",
            str(pool_path),
            "--output",
            str(output_path),
            "--academic-year",
            "1404",
            "--policy",
            str(policy_path),
        ]
    )

    assert rc == 0
    assert output_path.exists()
