from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.core.common.types import HeaderMode
from app.infra.cli_legacy import (
    _attach_sabt_sheet_if_selected,
    _build_sabt_allocations_if_needed,
)
from app.infra.io_utils import write_xlsx_atomic
from app.infra.excel import export_allocations as sabt_exports


def test_allocations_sabt_sheet_included_without_sabt_output(tmp_path: Path) -> None:
    profile_df = pd.DataFrame(
        {
            sabt_exports._HEADER_COLUMN: [
                "کد ثبت نام0",
                "پیدا کردن ردیف پشتیبان از فیلد 141",
            ],
            sabt_exports._VALUE_COLUMN: ["student_id", ""],
            sabt_exports._ORDER_COLUMN: [1, 2],
            sabt_exports._SOURCE_COLUMN: [
                sabt_exports._SOURCE_STUDENT,
                sabt_exports._SOURCE_ALLOCATION,
            ],
        }
    )
    profile_path = tmp_path / "profile.xlsx"
    profile_df.to_excel(
        profile_path, sheet_name=sabt_exports._PROFILE_SHEET_NAME, index=False
    )

    allocations_df = pd.DataFrame(
        {"__source_index__": [0], "student_id": ["S-1"], "mentor_id": [101]}
    )
    students_df = pd.DataFrame(
        {
            "__source_index__": [0],
            "student_id": ["S-1"],
            "student_first_name": ["Test"],
            "student_last_name": ["Student"],
        }
    )

    sabt_allocations_df = _build_sabt_allocations_if_needed(
        allocations_df=allocations_df,
        students_df=students_df,
        export_profile_choice="sabt",
        export_profile_path=str(profile_path),
        summary_df=None,
    )

    sheets: dict[str, pd.DataFrame] = {
        "allocations": allocations_df,
        "logs": pd.DataFrame({"student_id": ["S-1"], "allocation_status": ["success"]}),
        "updated_pool": pd.DataFrame({"mentor_id": [101], "remaining_capacity": [1]}),
    }
    header_overrides: dict[str, HeaderMode | None] = {}
    _attach_sabt_sheet_if_selected(
        sheets=sheets,
        header_overrides=header_overrides,
        export_profile_choice="sabt",
        sabt_allocations_df=sabt_allocations_df,
    )

    output_path = tmp_path / "allocations.xlsx"
    write_xlsx_atomic(
        sheets,
        output_path,
        header_mode=None,
        sheet_header_modes=header_overrides,
        sheet_prepare_modes={"allocations_sabt": "raw"},
    )

    workbook = load_workbook(output_path, read_only=True)
    assert "allocations_sabt" in workbook.sheetnames
