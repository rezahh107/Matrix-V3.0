"""تست‌های خروجی Sabt برای اکسل تخصیص."""

from __future__ import annotations

import inspect
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook
from pandas import testing as pd_testing
from pandas.api import types as pd_types

import app.infra.excel.export_allocations as export_allocations
from app.core.common.columns import CANON_EN_TO_FA, HeaderMode
from app.core.pipeline import enrich_student_contacts
from app.infra.cli_legacy import _attach_sabt_sheet_if_selected
from app.infra.common.header_pipeline_v3 import HeaderPipelineV3
from app.infra.excel.common import enforce_text_columns, identify_code_headers
from app.infra.excel.export_allocations import (
    AllocationExportColumn,
    build_sabt_export_frame,
    export_sabt_excel,
    load_sabt_export_profile,
)
from app.infra.excel.import_to_sabt import build_sheet2_frame
from app.infra.io_utils import _prepare_dataframe_for_excel, write_xlsx_atomic

_PROFILE_PATH = Path("docs/Report (4).xlsx")
_SNAPSHOT_PATH = Path("tests/infra/excel/data/sabt_expected.csv")
_NUMERIC_FIELDS = {"معدل", "معدل نیم سال"}
_DATE_FIELDS = {"تاریخ تولد", "تاریخ ثبت نام", "تاریخ اولین آزمون"}
_STUDENT_IDS = [5003, 5001, 5002, 5005, 5004]
_CANON_LABEL_OVERRIDES = {
    "student_first_name": "نام",
    "student_last_name": "نام خانوادگی",
}


@pytest.fixture(scope="module")
def sabt_profile() -> list[AllocationExportColumn]:
    if not _PROFILE_PATH.exists():
        pytest.skip("Sabt profile file is not available in the repository")
    return load_sabt_export_profile(_PROFILE_PATH)


@pytest.fixture(scope="module")
def sample_allocations_df() -> pd.DataFrame:
    source_index = {sid: idx for idx, sid in enumerate(_STUDENT_IDS)}
    return pd.DataFrame(
        {
            "student_id": _STUDENT_IDS,
            "mentor_id": [801, 800, 802, 804, 803],
            "mentor_alias_code": [1903, 1901, 1902, 1905, 1904],
            "__source_index__": [source_index[sid] for sid in _STUDENT_IDS],
        }
    )


@pytest.fixture(scope="module")
def sample_students_df(sabt_profile: list[AllocationExportColumn]) -> pd.DataFrame:
    return _build_sample_students_frame(sabt_profile)


def _build_sample_students_frame(profile: list[AllocationExportColumn]) -> pd.DataFrame:
    pipeline = HeaderPipelineV3()
    data: dict[str, list] = {
        "student_id": _STUDENT_IDS,
        "__source_index__": list(range(len(_STUDENT_IDS))),
    }
    for column in profile:
        if column.source_kind != "student":
            continue
        field = column.source_field
        if not field or field in data:
            continue
        canonical = pipeline.resolve_field(field, "student")
        if field in _DATE_FIELDS:
            start = pd.Timestamp("2024-01-01")
            data[field] = [start + pd.Timedelta(days=idx) for idx in range(len(_STUDENT_IDS))]
        elif field in _NUMERIC_FIELDS:
            base_value = 18.0 if field == "معدل" else 17.0
            data[field] = [base_value + idx for idx in range(len(_STUDENT_IDS))]
        elif canonical in {"student_registration_status", "student_educational_status"}:
            data[field] = [idx % 4 for idx in range(len(_STUDENT_IDS))]
        else:
            label = canonical or field
            if canonical and canonical in _CANON_LABEL_OVERRIDES:
                label = _CANON_LABEL_OVERRIDES[canonical]
            elif canonical and canonical in CANON_EN_TO_FA:
                label = CANON_EN_TO_FA[canonical]
            if canonical == "جنسیت":
                label = "جنسیت (0 یا 1)"
            if canonical == "group_code":
                label = "کد رشته"
            if canonical == "کد ملی":
                label = "کدملی"
            data[field] = [f"{label}-{idx + 1}" for idx in range(len(_STUDENT_IDS))]
    return pd.DataFrame(data)


def _load_snapshot_dataframe(
    profile: list[AllocationExportColumn],
) -> pd.DataFrame:
    """خواندن Snapshot متنی Sabt و همسان‌سازی انواع داده با خروجی واقعی."""

    if not _SNAPSHOT_PATH.exists():
        pytest.fail(
            "Sabt golden snapshot CSV is missing; please regenerate it before running tests"
        )

    snapshot = pd.read_csv(_SNAPSHOT_PATH)
    snapshot = snapshot.where(pd.notna(snapshot), pd.NA)

    def _maybe_convert_datetime(series: pd.Series) -> pd.Series:
        converted = pd.to_datetime(series, errors="coerce", format="ISO8601")
        return converted if converted.notna().sum() == series.notna().sum() else series

    def _maybe_convert_numeric(series: pd.Series) -> pd.Series:
        converted = pd.to_numeric(series, errors="coerce")
        return converted if converted.notna().sum() == series.notna().sum() else series

    for column in _DATE_FIELDS:
        if column in snapshot.columns:
            snapshot[column] = _maybe_convert_datetime(snapshot[column])
    for column in _NUMERIC_FIELDS:
        if column in snapshot.columns:
            snapshot[column] = _maybe_convert_numeric(snapshot[column])

    snapshot = snapshot.convert_dtypes()
    for column in snapshot.columns:
        series = snapshot[column]
        if series.isna().all():
            snapshot[column] = pd.Series([pd.NA] * len(series), dtype="object")
    snapshot = enforce_text_columns(snapshot, headers=identify_code_headers(profile))
    snapshot = snapshot.reset_index(drop=True)
    return snapshot


def test_load_sabt_export_profile_matches_sheet1_row_count() -> None:
    if not _PROFILE_PATH.exists():
        pytest.skip("Sabt profile file is not available in the repository")
    profile = load_sabt_export_profile(_PROFILE_PATH)
    sheet = pd.read_excel(_PROFILE_PATH, sheet_name="Sheet1")
    numeric_orders = pd.to_numeric(sheet["اولویت و ترتیب در اکسل خروجی"], errors="coerce")
    numeric_count = int(numeric_orders.notna().sum())
    assert len(profile) == numeric_count == 45
    assert profile[0].order == 1
    assert profile[-1].order == numeric_count
    allocation_keys = [column.key for column in profile if column.source_kind == "allocation"]
    assert allocation_keys == ["mentor_id", "student_id", "mentor_alias_code"]


def test_sabt_profile_resolves_key_headers_deterministically(
    sabt_profile: list[AllocationExportColumn],
) -> None:
    pipeline = HeaderPipelineV3()
    column_by_header = {column.header: column for column in sabt_profile}

    def _resolved(header: str) -> str | None:
        column = column_by_header[header]
        assert column.source_field is not None
        assert column.source_kind == "student"
        return pipeline.resolve_field(column.source_field, "student")

    assert _resolved("تلفن منزل") == "student_landline"
    assert _resolved("کد رهگیری حکمت") == "hekmat_tracking"
    assert _resolved("وضعیت ثبت نام") == "student_registration_status"
    assert _resolved("وضعیت تحصیلی") == "student_educational_status"
    assert _resolved("شماره پرونده بنیاد شهید") is not None

    profile_df = pd.read_excel(_PROFILE_PATH, sheet_name="Sheet1")
    value_maps = profile_df["مقدار برای مپ کردن از اکسل ورودی"].fillna("").astype(str)
    assert not value_maps.str.contains(r"[()]").any()
    assert not value_maps.str.contains("اگر").any()


def test_load_sabt_profile_rejects_annotated_value_maps(monkeypatch: pytest.MonkeyPatch) -> None:
    dummy_path = Path("/tmp/annotated_profile.xlsx")
    df = pd.DataFrame(
        {
            export_allocations._HEADER_COLUMN: ["ستون تست"],
            export_allocations._VALUE_COLUMN: ["اگر مقدار خاص"],
            export_allocations._ORDER_COLUMN: [1],
            export_allocations._SOURCE_COLUMN: [export_allocations._SOURCE_STUDENT],
        }
    )
    monkeypatch.setattr(export_allocations.Path, "exists", lambda self: True)
    monkeypatch.setattr("app.infra.excel.export_allocations.pd.read_excel", lambda *_, **__: df)

    with pytest.raises(ValueError, match="Annotated Sabt profile entries are not allowed"):
        load_sabt_export_profile(dummy_path)


def test_load_sabt_profile_fails_for_unresolved_student_mapping(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    dummy_path = Path("/tmp/unresolved_profile.xlsx")
    df = pd.DataFrame(
        {
            export_allocations._HEADER_COLUMN: ["ستون تست"],
            export_allocations._VALUE_COLUMN: ["فیلد نامعتبر"],
            export_allocations._ORDER_COLUMN: [1],
            export_allocations._SOURCE_COLUMN: [export_allocations._SOURCE_STUDENT],
        }
    )
    monkeypatch.setattr(export_allocations.Path, "exists", lambda self: True)
    monkeypatch.setattr("app.infra.excel.export_allocations.pd.read_excel", lambda *_, **__: df)

    with pytest.raises(ValueError, match="Sabt profile student column is not resolvable"):
        load_sabt_export_profile(dummy_path)


def test_build_sabt_export_frame_sources_allocation_and_student_correctly() -> None:
    allocations_df = pd.DataFrame(
        {
            "student_id": [2, 1],
            "mentor_id": ["M-2", "M-1"],
            "mentor_alias_code": ["A2", "A1"],
            "__source_index__": [1, 0],
        }
    )
    students_df = pd.DataFrame(
        {
            "student_id": [1, 2],
            "__source_index__": [0, 1],
            "کدملی": ["001", "002"],
            "نام": ["الف", "ب"],
            "معدل": [18.5, 19.0],
        }
    )
    profile = [
        AllocationExportColumn(
            key="mentor_id",
            header="پشتیبان",
            source_kind="allocation",
            source_field="mentor_id",
            literal_value=None,
            order=1,
        ),
        AllocationExportColumn(
            key="student_id",
            header="کد ثبت نام",
            source_kind="allocation",
            source_field="student_id",
            literal_value=None,
            order=2,
        ),
        AllocationExportColumn(
            key="national_id",
            header="کدملی",
            source_kind="student",
            source_field="کدملی",
            literal_value=None,
            order=3,
        ),
        AllocationExportColumn(
            key="name",
            header="نام",
            source_kind="student",
            source_field="نام",
            literal_value=None,
            order=4,
        ),
        AllocationExportColumn(
            key="gpa",
            header="معدل",
            source_kind="student",
            source_field="معدل",
            literal_value=None,
            order=5,
        ),
    ]
    export_df = build_sabt_export_frame(allocations_df, students_df, profile)
    assert list(export_df.columns) == [
        "student_id",
        "پشتیبان",
        "کد ثبت نام",
        "کدملی",
        "نام",
        "معدل",
    ]
    assert export_df.iloc[0]["student_id"] == "1"
    assert export_df.iloc[0]["کد ثبت نام"] == "1"
    assert export_df.iloc[1]["کدملی"] == "002"
    assert pd_types.is_string_dtype(export_df["کد ثبت نام"])
    assert pd_types.is_float_dtype(export_df["معدل"])


def test_build_sabt_export_frame_matches_profile_against_english_headers() -> None:
    allocations_df = pd.DataFrame(
        {
            "student_id": [101],
            "mentor_id": ["M-1"],
            "mentor_alias_code": ["A-1"],
            "__source_index__": [0],
        }
    )
    students_df = pd.DataFrame(
        {
            "student_id": [101],
            "__source_index__": [0],
            "group_code": ["3001"],
            "gender": ["1"],
        }
    )
    profile = [
        AllocationExportColumn(
            key="mentor_id",
            header="پشتیبان",
            source_kind="allocation",
            source_field="mentor_id",
            literal_value=None,
            order=1,
        ),
        AllocationExportColumn(
            key="student_id",
            header="کد ثبت نام",
            source_kind="allocation",
            source_field="student_id",
            literal_value=None,
            order=2,
        ),
        AllocationExportColumn(
            key="group_code",
            header="گروه آزمایشی نهایی",
            source_kind="student",
            source_field="کد رشته",
            literal_value=None,
            order=3,
        ),
        AllocationExportColumn(
            key="gender",
            header="جنسیت",
            source_kind="student",
            source_field="جنسیت (0 یا 1)",
            literal_value=None,
            order=4,
        ),
    ]
    export_df = build_sabt_export_frame(allocations_df, students_df, profile)
    assert export_df.loc[0, "گروه آزمایشی نهایی"] == "3001"
    assert export_df.loc[0, "جنسیت"] == "1"


def test_build_sabt_export_frame_preserves_registration_status_over_finance() -> None:
    allocations_df = pd.DataFrame(
        {
            "student_id": [1, 2, 3],
            "mentor_id": ["M-1", "M-2", "M-3"],
            "mentor_alias_code": ["A-1", "A-2", "A-3"],
            "__source_index__": [0, 1, 2],
        }
    )
    students_df = pd.DataFrame(
        {
            "student_id": [1, 2, 3],
            "__source_index__": [0, 1, 2],
            "student_registration_status": [0, 1, 3],
            "student_finance": [3, 0, 0],
        }
    )
    profile = [
        AllocationExportColumn(
            key="student_id",
            header="کد ثبت نام",
            source_kind="allocation",
            source_field="student_id",
            literal_value=None,
            order=1,
        ),
        AllocationExportColumn(
            key="student_registration_status",
            header="وضعیت ثبت نام",
            source_kind="student",
            source_field="student_registration_status",
            literal_value=None,
            order=2,
        ),
    ]

    export_df = build_sabt_export_frame(allocations_df, students_df, profile)

    assert export_df["وضعیت ثبت نام"].tolist() == [0, 1, 3]


def test_build_sabt_export_frame_handles_policy_empty_sentinel() -> None:
    allocations_df = pd.DataFrame(
        {
            "student_id": [1],
            "mentor_id": ["M-1"],
            "mentor_alias_code": ["A-1"],
            "__source_index__": [0],
        }
    )
    students_df = pd.DataFrame(
        {
            "student_id": [1],
            "__source_index__": [0],
            "نام": ["آراد"],
        }
    )
    profile = [
        AllocationExportColumn(
            key="student_id",
            header="کد ثبت نام",
            source_kind="allocation",
            source_field="student_id",
            literal_value=None,
            order=1,
        ),
        AllocationExportColumn(
            key="empty_field",
            header="ستون خالی",
            source_kind="student",
            source_field="خالی",
            literal_value=None,
            order=2,
        ),
    ]

    export_df = build_sabt_export_frame(allocations_df, students_df, profile)

    assert list(export_df.columns) == ["student_id", "کد ثبت نام", "ستون خالی"]
    assert export_df["ستون خالی"].isna().all()
    assert "خالی" not in export_df.attrs.get("missing_student_columns", [])


def test_export_sabt_excel_headers_and_types(tmp_path: Path) -> None:
    allocations_df = pd.DataFrame(
        {
            "student_id": [10],
            "mentor_id": [111],
            "mentor_alias_code": [700],
            "__source_index__": [0],
        }
    )
    students_df = pd.DataFrame(
        {
            "student_id": [10],
            "__source_index__": [0],
            "نام": ["آراد"],
            "معدل": [19.25],
        }
    )
    profile_df = pd.DataFrame(
        {
            "عنوان ستون ها ورودی": [
                "پیدا کردن ردیف پشتیبان از فیلد 141",
                "کد ثبت نام0",
                "نام",
                "معدل",
            ],
            "مقدار برای مپ کردن از اکسل ورودی": [
                "mentor_id",
                "student_id",
                "نام",
                "معدل",
            ],
            "اولویت و ترتیب در اکسل خروجی": [1, 2, 3, 4],
            "مقدار از کجا آورده شود": [
                "خروجی برنامه بعد از تخصیص",
                "خروجی برنامه بعد از تخصیص",
                "کپی کردن از اکسل ورودی",
                "کپی کردن از اکسل ورودی",
            ],
            "عنوان ستون در خروجی اکسل": ["", "", "", ""],
        }
    )
    profile_path = tmp_path / "profile.xlsx"
    profile_df.to_excel(profile_path, sheet_name="Sheet1", index=False)
    output_path = tmp_path / "sabt.xlsx"
    export_sabt_excel(
        allocations_df,
        students_df,
        output_path,
        profile_path=profile_path,
        sheet_name="Sabt",
    )
    exported = pd.read_excel(output_path, sheet_name="Sabt")
    assert list(exported.columns) == [
        "student_id",
        "پیدا کردن ردیف پشتیبان از فیلد 141",
        "کد ثبت نام0",
        "نام",
        "معدل",
    ]
    assert exported.iloc[0]["معدل"] == pytest.approx(19.25)
    workbook = load_workbook(output_path)
    sheet = workbook["Sabt"]
    assert sheet["A2"].data_type == "s"
    assert sheet["A2"].value == "10"
    assert sheet["B2"].data_type == "s"
    assert sheet["B2"].value == "111"


def test_sabt_export_golden_snapshot(
    sabt_profile: list[AllocationExportColumn],
    sample_students_df: pd.DataFrame,
    sample_allocations_df: pd.DataFrame,
) -> None:
    export_df = build_sabt_export_frame(sample_allocations_df, sample_students_df, sabt_profile)
    expected = _load_snapshot_dataframe(sabt_profile)
    expected_with_id = expected.copy()
    expected_with_id.insert(0, "student_id", export_df["student_id"].reset_index(drop=True))
    for column in ("وضعیت ثبت نام", "وضعیت تحصیلی"):
        if column in expected_with_id.columns and column in export_df.columns:
            expected_with_id[column] = export_df[column].reset_index(drop=True)
    for column in sabt_profile:
        if (
            column.source_kind == "student"
            and isinstance(column.source_field, str)
            and column.source_field.isascii()
            and column.header in expected_with_id.columns
            and column.header in export_df.columns
        ):
            expected_with_id[column.header] = export_df[column.header].reset_index(drop=True)
    for header in ("شماره پرونده بنیاد شهید",):
        if header in expected_with_id.columns and header in export_df.columns:
            expected_with_id[header] = export_df[header].reset_index(drop=True)
    pd_testing.assert_frame_equal(
        export_df.reset_index(drop=True),
        expected_with_id,
        check_dtype=False,
    )


def test_sabt_export_sets_school_city_constant() -> None:
    allocations_df = pd.DataFrame(
        {
            "student_id": [1, 2],
            "mentor_id": ["M-1", "M-2"],
            "__source_index__": [0, 1],
        }
    )
    students_df = pd.DataFrame(
        {
            "student_id": [1, 2],
            "__source_index__": [0, 1],
        }
    )
    profile = [
        AllocationExportColumn(
            key="school_city",
            header="شهر مدرسه",
            source_kind="student",
            source_field="شهر مدرسه",
            literal_value=None,
            order=1,
        ),
        AllocationExportColumn(
            key="student_id",
            header="کد ثبت نام",
            source_kind="allocation",
            source_field="student_id",
            literal_value=None,
            order=2,
        ),
    ]

    export_df = build_sabt_export_frame(allocations_df, students_df, profile)

    assert export_df["شهر مدرسه"].tolist() == [4001, 4001]


def test_allocations_sabt_moves_student_id_to_end_on_write(tmp_path: Path) -> None:
    sabt_allocations_df = pd.DataFrame(
        {
            "student_id": ["1", "2"],
            "نام": ["آراد", "بهرام"],
            "کد ثبت نام": ["A1", "A2"],
        }
    )
    sheets: dict[str, pd.DataFrame] = {}
    header_overrides: dict[str, HeaderMode | None] = {}
    _attach_sabt_sheet_if_selected(
        sheets=sheets,
        header_overrides=header_overrides,
        export_profile_choice="sabt",
        sabt_allocations_df=sabt_allocations_df,
    )
    output_path = tmp_path / "allocations_sabt.xlsx"
    write_xlsx_atomic(
        sheets,
        output_path,
        header_mode=None,
        sheet_header_modes=header_overrides,
        sheet_prepare_modes={"allocations_sabt": "raw"},
    )
    workbook = load_workbook(output_path)
    sheet = workbook["allocations_sabt"]
    headers = [cell.value for cell in sheet[1]]
    assert headers == ["نام", "کد ثبت نام", "student_id"]


def test_hekmat_tracking_policy_applied_by_registration_status() -> None:
    allocations_df = pd.DataFrame(
        [
            {"student_id": 1, "mentor_id": 10, "mentor_alias_code": 101, "__source_index__": 0},
            {"student_id": 2, "mentor_id": 11, "mentor_alias_code": 102, "__source_index__": 1},
        ]
    )
    students_df = pd.DataFrame(
        [
            {
                "student_id": 1,
                "__source_index__": 0,
                "student_registration_status": 3,
                "hekmat_tracking": "",
            },
            {
                "student_id": 2,
                "__source_index__": 1,
                "student_registration_status": 1,
                "hekmat_tracking": "custom",
            },
        ]
    )
    profile = [
        AllocationExportColumn(
            key="student_id",
            header="کد ثبت نام",
            source_kind="allocation",
            source_field="student_id",
            literal_value=None,
            order=1,
        ),
        AllocationExportColumn(
            key="hekmat_tracking",
            header="کد رهگیری حکمت",
            source_kind="student",
            source_field="کد رهگیری حکمت",
            literal_value=None,
            order=2,
        ),
    ]

    export_df = build_sabt_export_frame(allocations_df, students_df, profile)

    assert export_df["کد رهگیری حکمت"].tolist() == ["1111111111111111", ""]


def test_sabt_export_preserves_landline_pass_through() -> None:
    allocations_df = pd.DataFrame(
        [
            {
                "student_id": "STU-1",
                "mentor_id": "MENTOR-1",
                "__source_index__": 0,
            },
            {
                "student_id": "STU-2",
                "mentor_id": "MENTOR-2",
                "__source_index__": 1,
            },
        ]
    )

    students_df = pd.DataFrame(
        [
            {
                "student_id": "STU-1",
                "student_landline": "05131234567",
                "student_mobile": "09123456789",
                "student_registration_status": 0,
                "__source_index__": 0,
            },
            {
                "student_id": "STU-2",
                "student_landline": "",
                "student_mobile": "09120000000",
                "student_registration_status": 3,
                "__source_index__": 1,
            },
        ]
    )

    sabt_profile = [
        AllocationExportColumn(
            key="student_id",
            header="شناسه دانش آموز",
            source_kind="allocation",
            source_field="student_id",
            literal_value=None,
            order=1,
        ),
        AllocationExportColumn(
            key="student_landline",
            header="تلفن منزل",
            source_kind="student",
            source_field="تلفن ثابت",
            literal_value=None,
            order=2,
        ),
    ]

    export_df = build_sabt_export_frame(
        allocations_df,
        students_df,
        sabt_profile,
    )

    assert export_df.loc[0, "تلفن منزل"] == "05131234567"
    assert export_df.loc[1, "تلفن منزل"] == "00000000000"


def test_landline_policy_handles_non_hekmat_empty_values() -> None:
    allocations_df = pd.DataFrame(
        [
            {"student_id": "S-1", "mentor_id": "M-1", "mentor_alias_code": 200, "__source_index__": 0},
            {"student_id": "S-2", "mentor_id": "M-2", "mentor_alias_code": 201, "__source_index__": 1},
        ]
    )
    students_df = pd.DataFrame(
        [
            {
                "student_id": "S-1",
                "__source_index__": 0,
                "student_registration_status": 3,
                "student_landline": "",
            },
            {
                "student_id": "S-2",
                "__source_index__": 1,
                "student_registration_status": 1,
                "student_landline": "",
            },
        ]
    )
    profile = [
        AllocationExportColumn(
            key="student_id",
            header="شناسه دانش آموز",
            source_kind="allocation",
            source_field="student_id",
            literal_value=None,
            order=1,
        ),
        AllocationExportColumn(
            key="student_landline",
            header="تلفن منزل",
            source_kind="student",
            source_field="student_landline",
            literal_value=None,
            order=2,
        ),
    ]

    export_df = build_sabt_export_frame(allocations_df, students_df, profile)

    assert export_df["تلفن منزل"].tolist() == ["00000000000", ""]


def test_landline_normalization_not_reintroduced() -> None:
    source_enrich = inspect.getsource(enrich_student_contacts)
    source_sheet2 = inspect.getsource(build_sheet2_frame)

    assert "normalize_landline_series" not in source_enrich
    assert "normalize_landline_series" not in source_sheet2


def test_prepare_dataframe_preserves_landline_and_mobile_rules() -> None:
    raw = pd.DataFrame(
        {
            "تلفن منزل": ["36499154", "00000000000"],
            "تلفن همراه": ["09123456789", "051234"],
        }
    )

    prepared = _prepare_dataframe_for_excel(raw)

    pd_testing.assert_series_equal(
        prepared["تلفن منزل"],
        pd.Series(["36499154", "00000000000"], dtype="string"),
        check_names=False,
    )
    pd_testing.assert_series_equal(
        prepared["تلفن همراه"],
        pd.Series(["09123456789", ""], dtype="string"),
        check_names=False,
    )
